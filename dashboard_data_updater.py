#!/usr/bin/env python3
"""
从 UUAtO2 读取全部每日数据 + 从 KFTIUP 读取主播班次数据，供前端仪表盘使用

敏感配置通过环境变量传入（供 GitHub Actions 使用）：
  LARK_CLI                lark-cli 可执行文件路径（默认本机路径）
  DASHBOARD_SHEET_TOKEN   仪表盘数据源表 token（UUAtO2）
  HANDOFF_SHEET_TOKEN     主播班次表 token（KFTIUP）
身份固定使用 bot（--as bot），便于在 CI 中运行。

漏斗数据（funnel）从本地最新「直播明细_全部账号_*.xlsx」提取；CI 云端无
本地 Excel 时保留上一版值（不覆盖）。
"""
import json, subprocess, sys, os, re, ast, glob, argparse
from datetime import datetime, date

LARK_CLI = os.environ.get("LARK_CLI", "/Users/apple/Documents/Codex/2026-06-03/cli/lark-cli")
TOKEN = os.environ.get("DASHBOARD_SHEET_TOKEN", "Jos6sfYSRh4eWXtZalBcoFetnCe")
HANDOFF = os.environ.get("HANDOFF_SHEET_TOKEN", "shtcnwOdgFZCQAf4ZjiR5egkoTc")
HANDOFF_SHEET = "KFTIUP"
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard_data.json")
FOOD_DAILY = os.environ.get("FOOD_DAILY_SHEET_TOKEN", "CD2psfSlghdutdt9wHmcOJaDntf")
FOOD_SUMMARY_SHEET = "KYcREK"
FOOD_ANCHOR_SHEET = "cdzpqi"
FOOD_ANCHOR_LEGACY_SHEET = "sXKpGG"
FOOD_SHOP_SHEET = "9VVamg"
FOOD_LIVE_SHEET = "weIxvN"

def lark_read(token, range_expr, value_render=None, identity="bot", timeout=15):
    cmd = [LARK_CLI, "sheets", "+read", "--as", identity,
           "--spreadsheet-token", token,
           "--range", range_expr,
           "--format", "json"]
    if value_render:
        cmd += ["--value-render-option", value_render]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        data = json.loads(r.stdout)
        if data.get("ok") and "data" in data and "valueRange" in data["data"]:
            return data["data"]["valueRange"]["values"]
        else:
            print(f"[warn] lark read failed: {data.get('error', r.stderr[:200])}", file=sys.stderr)
            return None
    except Exception as e:
        print(f"[error] {e}", file=sys.stderr)
        return None

def pn(v):
    if v is None or v == "" or v == " ":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except:
        return None

def col_num(index):
    """0-based column index to letter(s)."""
    result = ""
    n = index + 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result

def eval_formula(value, cells, visiting=None):
    """Evaluate formula strings like '30001-C3' or 'C3/C4' using resolved cell values."""
    if not isinstance(value, str):
        return pn(value) or 0.0
    expr = value.strip()
    if not expr:
        return 0.0
    if visiting is None:
        visiting = set()

    def replace_ref(match):
        ref = match.group(0)
        # Skip cross-sheet refs like '阴山支援主播业绩表格'!V29
        if "!" in ref:
            return "0"
        if ref in visiting:
            return "0"
        visiting.add(ref)
        resolved = eval_formula(cells.get(ref), cells, visiting)
        visiting.remove(ref)
        return str(resolved)

    expr = re.sub(r"\b[A-Z]{1,3}[0-9]+\b", replace_ref, expr)
    if not re.fullmatch(r"[0-9+\-*/(). ]+", expr):
        return 0.0
    try:
        tree = ast.parse(expr, mode="eval")
        for node in ast.walk(tree):
            if not isinstance(node, (
                ast.Expression, ast.BinOp, ast.UnaryOp,
                ast.Add, ast.Sub, ast.Mult, ast.Div,
                ast.USub, ast.UAdd, ast.Constant, ast.Load,
            )):
                return 0.0
        return float(eval(compile(tree, "<formula>", "eval"), {"__builtins__": {}}, {}))
    except:
        return 0.0

def excel_serial(day):
    """date -> Excel serial number."""
    from datetime import date
    return (day - date(1899, 12, 30)).days


def serial_to_date(serial):
    """Excel serial number -> date."""
    from datetime import date
    return date(1899, 12, 30) + __import__("datetime").timedelta(days=int(serial))


def parse_anchor_block(values, start_idx):
    """Parse one 5-row anchor block starting at start_idx. Returns list of anchors."""
    anchors = []
    for col_offset in (1, 3, 5, 7):  # B, D, F, H
        name_row = values[start_idx]
        if len(name_row) <= col_offset + 1:
            continue
        name = name_row[col_offset + 1]
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip().replace("冲饮主播：", "").strip()
        if not name_str or name_str.startswith("主播"):
            continue

        hrs = 0
        if start_idx + 1 < len(values):
            hrs_row = values[start_idx + 1]
            if len(hrs_row) > col_offset + 1:
                hrs = pn(hrs_row[col_offset + 1]) or 0

        gmv = 0
        if start_idx + 2 < len(values):
            gmv_row = values[start_idx + 2]
            if len(gmv_row) > col_offset + 1:
                gmv = rounded(pn(gmv_row[col_offset + 1]) or 0, 2)

        cost = 0
        if start_idx + 3 < len(values):
            cost_row = values[start_idx + 3]
            if len(cost_row) > col_offset + 1:
                cost = rounded(pn(cost_row[col_offset + 1]) or 0, 2)

        if gmv == 0 and cost == 0:
            continue  # 无数据的空主播位跳过

        anchors.append({
            "name": name_str,
            "gmv": gmv,
            "spend": cost,
            "roi": round(gmv / cost, 2) if cost > 0 else 0,
            "hours": hrs,
        })
    return anchors


def fetch_anchor_blocks():
    """读取 KFTIUP 全部日期块（FormattedValue 直接取公式计算值）。
    返回 {serial: anchors_list}，按日期升序。
    A 列在 FormattedValue 模式下可能是日期字符串（如 2026/03/15）或 serial 数字，均需兼容。"""
    values = lark_read(HANDOFF, f"{HANDOFF_SHEET}!A:I", value_render="FormattedValue")
    if not values:
        return {}

    def to_serial(v):
        """把 A 列值转成 Excel serial。支持数字或日期字符串。"""
        n = pn(v)
        if n is not None and n > 0:
            return int(n)
        if isinstance(v, str) and "/" in v:
            try:
                import datetime as _dt
                d = _dt.datetime.strptime(v.strip(), "%Y/%m/%d").date()
                return excel_serial(d)
            except Exception:
                return None
        return None

    blocks = {}
    i = 0
    while i < len(values):
        row = values[i]
        serial = to_serial(row[0]) if row else None
        if serial is not None and serial > 0 and i + 4 < len(values):
            anchors = parse_anchor_block(values, i)
            if anchors:
                blocks[serial] = anchors
            i += 5
        else:
            i += 1
    return blocks


def fetch_anchors(target_serial=None):
    """读取指定日期（Excel serial）的主播数据。默认取 blocks 中 <= 最新一天的最近块。
    注意：KFTIUP 中可能存在"当日"（数据未填完）块，因此默认按调用方传入的
    target_serial（通常为最新完整数据日 = 前一日）取数，避免取到当日不完整数据。"""
    blocks = fetch_anchor_blocks()
    if not blocks:
        return []
    if target_serial is not None and int(target_serial) in blocks:
        return blocks[int(target_serial)]
    # 回退：取所有块中 serial 最大的
    last = max(blocks.keys())
    return blocks[last]


def fetch_anchor_history():
    """全部日期块的主播历史数据，转成 [{serial, date, anchors:[...]}] 升序。
    date 为 YYYY-MM-DD 字符串，供前端历史查询按日期范围筛选。"""
    blocks = fetch_anchor_blocks()
    if not blocks:
        return []
    history = []
    for serial in sorted(blocks.keys()):
        day = serial_to_date(serial)
        history.append({
            "serial": serial,
            "date": day.isoformat(),
            "anchors": blocks[serial],
        })
    return history


def build_funnel_from_workbook(path):
    """从指定直播明细 Excel 提取漏斗数据；无有效曝光时返回 None。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["直播间明细"]
        f = {
            "exposure": 0.0, "views": 0.0, "prod_exposure": 0.0, "prod_click": 0.0, "buyers": 0.0,
            "exposure_count": 0.0, "watch_count": 0.0, "comments": 0.0,
            "new_live_group": 0.0, "new_fans": 0.0, "new_shopping_group": 0.0,
            "avg_stay_min": 0.0,
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] == "阴山优麦冲饮旗舰店":
                f["exposure"] += float(row[6] or 0)
                f["exposure_count"] += float(row[7] or 0)
                f["views"] += float(row[8] or 0)
                f["watch_count"] += float(row[10] or 0)
                f["prod_exposure"] += float(row[20] or 0)
                f["prod_click"] += float(row[21] or 0)
                f["buyers"] += float(row[29] or 0)
                f["comments"] += float(row[14] or 0)
                f["new_live_group"] += float(row[15] or 0)
                f["new_fans"] += float(row[16] or 0)
                f["new_shopping_group"] += float(row[42] or 0)
                f["avg_stay_min"] = max(f["avg_stay_min"], float(row[13] or 0))
        wb.close()
        if f["exposure"] <= 0:
            return None
        def rate(a, b):
            return round(a / b * 100, 2) if b else 0
        import re as _re
        m = _re.search(r"(\d{8})_", path)
        return {
            **f,
            "rate_view": rate(f["views"], f["exposure"]),
            "rate_prod_exposure": rate(f["prod_exposure"], f["views"]),
            "rate_click": rate(f["prod_click"], f["prod_exposure"]),
            "rate_buy": rate(f["buyers"], f["prod_click"]),
            "rate_total": rate(f["buyers"], f["exposure"]),
            "rate_watch_buy": rate(f["buyers"], f["watch_count"]),
            "interactions": round(f["comments"] + f["new_live_group"] + f["new_fans"] + f["new_shopping_group"], 2),
            "rate_watch_view_times": rate(f["watch_count"], f["exposure_count"]),
            "rate_watch_interaction": rate(f["comments"] + f["new_live_group"] + f["new_fans"] + f["new_shopping_group"], f["watch_count"]),
            "rate_exposure_interaction": rate(f["comments"] + f["new_live_group"] + f["new_fans"] + f["new_shopping_group"], f["exposure_count"]),
            "date": m.group(1) if m else "",
        }
    except Exception as e:
        print(f"[warn] funnel excel: {e}", file=sys.stderr)
        return None


def fetch_funnel_from_excel():
    """取本地最新一份直播明细 Excel 的漏斗数据。"""
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        files = sorted(glob.glob(os.path.join(base_dir, "直播明细_全部账号_*.xlsx")))
        if not files:
            return None
        return build_funnel_from_workbook(files[-1])
    except Exception as e:
        print(f"[warn] funnel excel: {e}", file=sys.stderr)
        return None


def fetch_prev_funnel_from_excel():
    """取本地倒数第二份直播明细 Excel 的昨日漏斗数据。"""
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        files = sorted(glob.glob(os.path.join(base_dir, "直播明细_全部账号_*.xlsx")))
        if len(files) < 2:
            return None
        return build_funnel_from_workbook(files[-2])
    except Exception as e:
        print(f"[warn] prev funnel excel: {e}", file=sys.stderr)
        return None


def parse_live_date(value):
    """cC79qR 的开始时间可能是 Excel 序列号或 YYYY/M/D H:MM 字符串。"""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return serial_to_date(int(value)).isoformat()
    s = str(value).strip()
    m = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    return None

def fetch_live_history():
    """从 cC79qR 读取阴山优麦冲饮旗舰店每日直播明细，供历史漏斗按日期区间聚合。"""
    values = lark_read(TOKEN, "cC79qR!A1:BB200")
    if not values:
        return []
    records = []
    for row in values[1:]:
        if not row or len(row) < 44:
            continue
        if str(row[1] or "").strip() != "阴山优麦冲饮旗舰店":
            continue
        dstr = parse_live_date(row[3])
        if not dstr:
            continue
        records.append({
            "date": dstr,
            "exposure": pn(row[6]) or 0,
            "exposure_count": pn(row[7]) or 0,
            "views": pn(row[8]) or 0,
            "watch_count": pn(row[10]) or 0,
            "avg_stay_min": pn(row[13]) or 0,
            "comments": pn(row[14]) or 0,
            "new_live_group": pn(row[15]) or 0,
            "new_fans": pn(row[16]) or 0,
            "prod_exposure": pn(row[20]) or 0,
            "prod_click": pn(row[21]) or 0,
            "gmv": pn(row[25]) or 0,
            "hourly_gmv": pn(row[27]) or 0,
            "buyers": pn(row[29]) or 0,
            "refund_amount": pn(row[31]) or 0,
            "new_shopping_group": pn(row[42]) or 0,
            "cost": pn(row[43]) or 0,
        })
    records.sort(key=lambda r: r["date"])
    return records

def rounded(v, d=2):
    return round(float(v) + 1e-9, d)

def default_food():
    return {
        "knownDays": [],
        "latest": {"day": 0, "b": None, "c": None, "d": None, "e": None, "f": None, "g": None, "h": None, "date": None},
        "targetGmv": 840000,
        "targetGsv": 720000,
        "monthly": {"gmv": 0, "spend": 0, "gsv": 0, "duration": 0, "roi": 0},
        "funnel": None,
        "funnelPrev": None,
        "liveHistory": [],
        "anchors": [],
        "anchorHistory": [],
    }

def fetch_food_known_days():
    """从 9VVamg 读取食品店 1-8 月逐日成交数据。"""
    values = lark_read(TOKEN, f"{FOOD_SHOP_SHEET}!A2:I245", identity="user", timeout=30)
    if not values:
        return []
    out = []
    for row in values:
        if not row:
            continue
        serial = pn(row[0])
        dv = pn(row[1]) if len(row) > 1 else None
        if serial is None or serial <= 40000 or dv is None:
            continue
        d = serial_to_date(int(serial))
        out.append({
            "date": d.isoformat(),
            "day": d.day,
            "dv": dv,
            "sp": pn(row[2]) if len(row) > 2 else None,
            "duration": None,
            "gsv": None,
            "hourly": None,
            "roi": pn(row[8]) if len(row) > 8 else None,
            "live": pn(row[3]) if len(row) > 3 else 0,
            "short_video": pn(row[4]) if len(row) > 4 else 0,
            "card": pn(row[5]) if len(row) > 5 else 0,
            "other": pn(row[6]) if len(row) > 6 else 0,
            "graphic": pn(row[7]) if len(row) > 7 else 0,
        })
    return out

def fetch_food_summary():
    """从 KYcREK 读取食品店每日消耗/时长/GSV 与目标。"""
    values = lark_read(FOOD_DAILY, f"{FOOD_SUMMARY_SHEET}!A1:J45", value_render="FormattedValue", identity="user", timeout=30)
    if not values:
        return {}, {}
    summary = {}
    for row in values[3:34]:
        if len(row) < 10:
            continue
        ds = str(row[1] or "").strip()
        m = re.match(r"(\d+)月(\d+)日", ds)
        if not m:
            continue
        date_str = f"2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        summary[date_str] = {
            "sp": pn(row[5]),
            "duration": pn(row[3]),
            "hourly": pn(row[4]),
            "gsv": pn(row[8]),
            "roi": pn(row[9]),
        }
    targets = {}
    if len(values) > 36:
        targets["targetGmv"] = pn(values[36][2]) if len(values[36]) > 2 else None
        targets["targetGsv"] = pn(values[36][8]) if len(values[36]) > 8 else None
        targets["gapGmv"] = pn(values[37][2]) if len(values) > 37 and len(values[37]) > 2 else None
        targets["gapGsv"] = pn(values[37][8]) if len(values) > 37 and len(values[37]) > 8 else None
    return summary, targets

def fetch_food_live_history():
    """从 weIxvN 读取阴山优麦食品旗舰店直播明细。"""
    values = lark_read(TOKEN, f"{FOOD_LIVE_SHEET}!A2:BB199", identity="user", timeout=30)
    if not values:
        return []
    records = []
    for row in values:
        if not row or len(row) < 44:
            continue
        if str(row[1] or "").strip() != "阴山优麦食品旗舰店":
            continue
        dstr = parse_live_date(row[3])
        if not dstr:
            continue
        records.append({
            "date": dstr,
            "duration": pn(row[5]) / 60 if pn(row[5]) else 0,
            "exposure": pn(row[6]) or 0,
            "exposure_count": pn(row[7]) or 0,
            "views": pn(row[8]) or 0,
            "watch_count": pn(row[10]) or 0,
            "avg_stay_min": pn(row[13]) or 0,
            "comments": pn(row[14]) or 0,
            "new_live_group": pn(row[15]) or 0,
            "new_fans": pn(row[16]) or 0,
            "prod_exposure": pn(row[20]) or 0,
            "prod_click": pn(row[21]) or 0,
            "gmv": pn(row[25]) or 0,
            "hourly_gmv": pn(row[27]) or 0,
            "buyers": pn(row[29]) or 0,
            "refund_amount": pn(row[31]) or 0,
            "new_shopping_group": pn(row[42]) or 0,
            "cost": pn(row[43]) or 0,
        })
    records.sort(key=lambda r: r["date"])
    return records

def build_food_funnel(records):
    if not records:
        return None
    s = {
        "exposure": 0.0, "exposure_count": 0.0, "views": 0.0, "watch_count": 0.0,
        "prod_exposure": 0.0, "prod_click": 0.0, "buyers": 0.0, "comments": 0.0,
        "new_live_group": 0.0, "new_fans": 0.0, "new_shopping_group": 0.0, "avg_stay_min": 0.0,
    }
    for r in records:
        for k in s:
            s[k] += float(r.get(k) or 0)
    if s["exposure"] <= 0:
        return None
    def rate(a, b):
        return round(a / b * 100, 2) if b else 0
    return {
        **s,
        "rate_view": rate(s["views"], s["exposure"]),
        "rate_prod_exposure": rate(s["prod_exposure"], s["views"]),
        "rate_click": rate(s["prod_click"], s["prod_exposure"]),
        "rate_buy": rate(s["buyers"], s["prod_click"]),
        "rate_total": rate(s["buyers"], s["exposure"]),
        "rate_watch_buy": rate(s["buyers"], s["watch_count"]),
        "interactions": round(s["comments"] + s["new_live_group"] + s["new_fans"] + s["new_shopping_group"], 2),
        "rate_watch_view_times": rate(s["watch_count"], s["exposure_count"]),
        "rate_watch_interaction": rate(s["comments"] + s["new_live_group"] + s["new_fans"] + s["new_shopping_group"], s["watch_count"]),
        "rate_exposure_interaction": rate(s["comments"] + s["new_live_group"] + s["new_fans"] + s["new_shopping_group"], s["exposure_count"]),
        "date": records[0]["date"].replace("-", ""),
    }

def parse_food_anchor_name(header):
    m = re.search(r"主播\s*\n+\s*([^\n]+)", str(header or ""))
    if m:
        return m.group(1).strip()
    text = str(header or "").replace("主播", "").strip()
    return text.splitlines()[0].strip() if text.splitlines() else ""

def fetch_food_anchor_blocks_cdzpqi():
    """从 cdzpqi 交班数据读取食品店主播班次块（优先）。"""
    values = lark_read(FOOD_DAILY, f"{FOOD_ANCHOR_SHEET}!A1:I500", identity="user", timeout=30)
    if not values:
        return {}
    blocks = {}
    i, n = 0, len(values)
    while i < n:
        row = values[i]
        serial = pn(row[0]) if row else None
        if serial is None or serial <= 40000:
            i += 1
            continue
        anchors = []
        if i + 4 < n:
            for idx in (2, 4, 6, 8):
                name = str(values[i][idx]).strip() if len(values[i]) > idx and values[i][idx] else ''
                gmv = pn(values[i + 2][idx]) if len(values[i + 2]) > idx else 0
                cost = pn(values[i + 3][idx]) if len(values[i + 3]) > idx else 0
                hours = pn(values[i + 1][idx]) if len(values[i + 1]) > idx else 0
                if not name:
                    continue
                if gmv or cost:
                    anchors.append({
                        "name": name,
                        "gmv": rounded(gmv or 0),
                        "spend": rounded(cost or 0),
                        "roi": rounded((gmv or 0) / cost, 2) if cost else 0,
                        "hours": rounded(hours or 0, 2),
                    })
        if anchors:
            blocks[int(serial)] = anchors
        i += 5
    return blocks


def fetch_food_anchor_blocks_legacy():
    """从 sXKpGG 读取食品店主播班次块（历史数据）。"""
    values = lark_read(FOOD_DAILY, f"{FOOD_ANCHOR_LEGACY_SHEET}!A1:G3181", identity="user", timeout=60)
    if not values:
        return {}
    blocks = {}
    i, n = 0, len(values)
    while i < n:
        row = values[i]
        serial = pn(row[0]) if row else None
        if serial is None or serial <= 40000:
            i += 1
            continue
        anchors = []
        j = i + 1
        while j < n:
            rowj = values[j]
            a0 = rowj[0] if rowj else None
            a0n = pn(a0)
            if a0 is not None and a0n is not None and a0n > 40000:
                break
            header = str(a0 or "")
            if "主播" not in header or "核心指标" in header:
                j += 1
                continue
            name = parse_food_anchor_name(header)
            gmv = pn(values[j + 1][4]) if j + 1 < n and len(values[j + 1]) > 4 else 0
            roi = pn(values[j + 2][4]) if j + 2 < n and len(values[j + 2]) > 4 else 0
            cost = pn(values[j + 5][3]) if j + 5 < n and len(values[j + 5]) > 3 else 0
            if name and (gmv or cost):
                anchors.append({
                    "name": name,
                    "gmv": rounded(gmv or 0),
                    "spend": rounded(cost or 0),
                    "roi": rounded(roi) if roi else (rounded((gmv or 0) / cost, 2) if cost else 0),
                    "hours": 0,
                })
            j += 6
        if anchors:
            blocks[int(serial)] = anchors
        i = j if j > i else i + 1
    return blocks

def fetch_food_anchor_history(blocks):
    return [{"serial": s, "date": serial_to_date(s).isoformat(), "anchors": blocks[s]} for s in sorted(blocks)]

def fetch_food_data():
    known = fetch_food_known_days()
    summary, targets = fetch_food_summary()
    by_date = {d["date"]: d for d in known}
    for ds, sm in summary.items():
        if ds in by_date:
            for k, v in sm.items():
                if by_date[ds].get(k) is None and v is not None:
                    by_date[ds][k] = v
    known = sorted(by_date.values(), key=lambda d: d["date"])
    live = fetch_food_live_history()
    blocks = fetch_food_anchor_blocks_cdzpqi()
    for serial, anchors in fetch_food_anchor_blocks_legacy().items():
        blocks.setdefault(serial, anchors)
    live_duration = {}
    for r in live:
        live_duration[r["date"]] = live_duration.get(r["date"], 0.0) + float(r.get("duration") or 0)
    for ds, d in by_date.items():
        if d.get("duration") is None and ds in live_duration:
            d["duration"] = round(live_duration[ds], 2)
        if d.get("hourly") is None and d.get("duration"):
            d["hourly"] = round((d.get("dv") or 0) / d["duration"], 2)
    latest = known[-1] if known else None
    latest_serial = excel_serial(date.fromisoformat(latest["date"])) if latest else None
    anchors = blocks.get(latest_serial, []) if latest_serial else []
    funnel = funnel_prev = None
    if live:
        dates = sorted(set(r["date"] for r in live))
        if dates:
            funnel = build_food_funnel([r for r in live if r["date"] == dates[-1]])
            if len(dates) > 1:
                funnel_prev = build_food_funnel([r for r in live if r["date"] == dates[-2]])
    aug = [d for d in known if d["date"].startswith("2026-08")]
    monthly = {
        "gmv": rounded(sum(d["dv"] for d in aug), 2),
        "spend": rounded(sum(d["sp"] or 0 for d in aug), 2),
        "gsv": rounded(sum(d["gsv"] or 0 for d in aug), 2),
        "duration": rounded(sum(d["duration"] or 0 for d in aug), 2),
        "roi": rounded(sum(d["dv"] for d in aug) / sum(d["sp"] or 0 for d in aug), 2) if sum(d["sp"] or 0 for d in aug) else 0,
    }
    latest_obj = {
        "date": latest["date"], "day": latest["day"], "b": latest["dv"], "c": latest["sp"],
        "d": latest["live"], "e": latest["short_video"], "f": latest["card"], "g": latest["other"], "h": latest["graphic"],
        "duration": latest["duration"], "hourly": latest["hourly"], "gsv": latest["gsv"], "roi": latest["roi"],
    } if latest else default_food()["latest"]
    return {
        "knownDays": known,
        "latest": latest_obj,
        "targetGmv": targets.get("targetGmv") or 840000,
        "targetGsv": targets.get("targetGsv") or 720000,
        "gapGmv": targets.get("gapGmv"),
        "gapGsv": targets.get("gapGsv"),
        "monthly": monthly,
        "funnel": funnel,
        "funnelPrev": funnel_prev,
        "liveHistory": live,
        "anchors": anchors,
        "anchorHistory": fetch_food_anchor_history(blocks),
    }

def fetch_data():
    rows = lark_read(TOKEN, "UUAtO2!A2:O32")
    if rows is None:
        return None

    out = {}

    # 所有已知日期数据
    known = []
    postal_hist = []
    latest_idx = None
    for i, row in enumerate(rows):
        dv = pn(row[1]) if len(row) > 1 else None
        sp = pn(row[2]) if len(row) > 2 else None
        if dv is not None:
            day_num = i + 1
            known.append({
                "day": day_num,
                "dv": dv,
                "sp": sp or 0,
                "live": pn(row[3]) if len(row) > 3 else None,
                "short_video": pn(row[4]) if len(row) > 4 else None,
                "card": pn(row[5]) if len(row) > 5 else None,
                "other": pn(row[6]) if len(row) > 6 else None,
                "graphic": pn(row[7]) if len(row) > 7 else None,
            })
            latest_idx = i

            pg = pn(row[9]) if len(row) > 9 else None
            ps = pn(row[10]) if len(row) > 10 else None
            if pg is not None and pg > 0:
                postal_hist.append({"day": day_num, "pg": pg, "ps": ps or 0})

    out["knownDays"] = known
    out["postalHistory"] = postal_hist

    # 最新一天的明细
    if latest_idx is not None:
        row = rows[latest_idx]
        out["latest"] = {
            "day": latest_idx + 1,
            "b": pn(row[1]) if len(row) > 1 else None,
            "c": pn(row[2]) if len(row) > 2 else None,
            "d": pn(row[3]) if len(row) > 3 else None,
            "e": pn(row[4]) if len(row) > 4 else None,
            "f": pn(row[5]) if len(row) > 5 else None,
            "g": pn(row[6]) if len(row) > 6 else None,
            "h": pn(row[7]) if len(row) > 7 else None,
        }
        pg = pn(row[9]) if len(row) > 9 else None
        ps = pn(row[10]) if len(row) > 10 else None
        out["postal"] = {
            "gmv": pg if pg else 0,
            "spend": ps if ps else 0,
            "roi": round(pg / ps, 2) if (pg is not None and ps and ps > 0) else 0
        }
        vg = pn(row[12]) if len(row) > 12 else None
        vs = pn(row[13]) if len(row) > 13 else None
        out["video"] = {
            "gmv": vg if vg else 0,
            "spend": vs if vs else 0,
            "roi": round(vg / vs, 2) if (vg is not None and vs and vs > 0) else 0
        }
    else:
        out["latest"] = {"day": 0}
        out["postal"] = {"gmv": 0, "spend": 0, "roi": 0}
        out["video"] = {"gmv": 0, "spend": 0, "roi": 0}

    # 主播班次数据（从 KFTIUP 读取）
    # target_serial = 最新完整数据日的 Excel serial（取 knownDays 最新一天对应日期）
    # 需求：主播栏维持前一日数据，不取 KFTIUP 中"当日"（可能未填完）的块
    if latest_idx is not None:
        latest_date_serial = pn(rows[latest_idx][0])
    else:
        latest_date_serial = None
    out["anchors"] = fetch_anchors(latest_date_serial)

    # 历史主播数据（全部日期块，供历史查询页使用）
    out["anchorHistory"] = fetch_anchor_history()

    # 读取月度汇总数据 from specific cells
    cmd2 = [LARK_CLI, "sheets", "+read", "--as", "bot",
           "--spreadsheet-token", TOKEN,
           "--range", "UUAtO2!J38:O38",
           "--value-render-option", "UnformattedValue",
           "--format", "json"]
    try:
        r2 = subprocess.run(cmd2, capture_output=True, text=True, timeout=15)
        d2 = json.loads(r2.stdout)
        if d2.get("ok") and "data" in d2 and "valueRange" in d2["data"]:
            sr = d2["data"]["valueRange"]["values"][0]
            out["postalMonthly"] = {
                "gmv": pn(sr[0]) if len(sr) > 0 and sr[0] is not None else 0,
                "spend": pn(sr[1]) if len(sr) > 1 and sr[1] is not None else 0,
                "roi": pn(sr[2]) if len(sr) > 2 and sr[2] is not None else 0,
            }
            out["videoMonthly"] = {
                "gmv": pn(sr[3]) if len(sr) > 3 and sr[3] is not None else 0,
                "spend": pn(sr[4]) if len(sr) > 4 and sr[4] is not None else 0,
                "roi": pn(sr[5]) if len(sr) > 5 and sr[5] is not None else 0,
            }
        else:
            out["videoMonthly"] = {"gmv": 0, "spend": 0, "roi": 0}
            out["postalMonthly"] = {"gmv": 0, "spend": 0, "roi": 0}
    except Exception as e:
        print(f"[warn] summary cells: {e}", file=sys.stderr)
        out["videoMonthly"] = {"gmv": 0, "spend": 0, "roi": 0}
        out["postalMonthly"] = {"gmv": 0, "spend": 0, "roi": 0}
    out["targetGmv"] = 2800000
    out["targetRoi"] = 3
    # 漏斗数据：优先本地 Excel（最新日期）；CI 云端无 Excel 时保留上一版值
    funnel = fetch_funnel_from_excel()
    funnel_prev = fetch_prev_funnel_from_excel()
    if funnel is not None:
        out["funnel"] = funnel
        if funnel_prev is not None:
            out["funnelPrev"] = funnel_prev
    elif os.path.exists(OUTPUT):
        try:
            prev = json.load(open(OUTPUT, encoding="utf-8"))
            if prev.get("funnel"):
                out["funnel"] = prev["funnel"]
            if prev.get("funnelPrev"):
                out["funnelPrev"] = prev["funnelPrev"]
        except Exception:
            pass
    if "funnel" not in out:
        out["funnel"] = None
    if "funnelPrev" not in out:
        out["funnelPrev"] = None
    # 历史直播明细（cC79qR），供历史查询页按日期区间聚合
    live_history = fetch_live_history()
    if live_history:
        out["liveHistory"] = live_history
    elif os.path.exists(OUTPUT):
        try:
            prev = json.load(open(OUTPUT, encoding="utf-8"))
            out["liveHistory"] = prev.get("liveHistory", [])
        except Exception:
            out["liveHistory"] = []
    else:
        out["liveHistory"] = []
    out["fetchedAt"] = subprocess.run(
        ["date", "+%Y-%m-%d %H:%M:%S"], capture_output=True, text=True
    ).stdout.strip()
    try:
        food = fetch_food_data()
    except Exception as e:
        print(f"[warn] food data: {e}", file=sys.stderr)
        food = None
    if food and (food.get("knownDays") or food.get("liveHistory") or food.get("anchorHistory")):
        out["food"] = food
    elif os.path.exists(OUTPUT):
        try:
            out["food"] = json.load(open(OUTPUT, encoding="utf-8")).get("food") or default_food()
        except Exception:
            out["food"] = default_food()
    else:
        out["food"] = default_food()

    return out

def deploy_dashboard():
    """提交并推送 dashboard_data.json，GitHub Pages 会随之更新。"""
    base = os.path.dirname(os.path.abspath(__file__))
    subprocess.run(["git", "-C", base, "add", "dashboard_data.json"], check=True)
    r = subprocess.run(["git", "-C", base, "diff", "--cached", "--quiet"], capture_output=True)
    if r.returncode == 0:
        print("[dashboard_data_updater] No dashboard data changes, skip push")
        return
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    subprocess.run(["git", "-C", base, "commit", "-m", f"chore: refresh dashboard data [{stamp}]"], check=True)
    subprocess.run(["git", "-C", base, "push"], check=True)
    print("[dashboard_data_updater] Pushed dashboard_data.json to origin/main")

def main():
    parser = argparse.ArgumentParser(description="Refresh dashboard data")
    parser.add_argument("--deploy", action="store_true", help="Commit and push dashboard_data.json to update GitHub Pages")
    args = parser.parse_args()
    print("[dashboard_data_updater] Fetching data from Feishu (UUAtO2 + KFTIUP)...")
    data = fetch_data()
    if data is None:
        print("[dashboard_data_updater] FAILED to fetch data", file=sys.stderr)
        sys.exit(1)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[dashboard_data_updater] Written {OUTPUT}")
    ld = data.get("latest", {})
    anc = data.get("anchors", [])
    print(f"[dashboard_data_updater] Latest day: {ld.get('day')}, GMV: {ld.get('b')}")
    print(f"[dashboard_data_updater] Anchors: {len(anc)} hosts")
    for a in anc:
        print(f"  {a['name']}: GMV {a['gmv']}, Cost {a['spend']}, ROI {a['roi']}")
    food = data.get("food") or {}
    fd = food.get("latest") or {}
    print(f"[dashboard_data_updater] Food latest: {fd.get('date')}, GMV: {fd.get('b')}, days: {len(food.get('knownDays') or [])}")
    if args.deploy:
        deploy_dashboard()

if __name__ == "__main__":
    main()
