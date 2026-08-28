#!/usr/bin/env python3
import argparse
import ast
import datetime as dt
import json
import re
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("/Users/apple/Desktop/日报文件夹")
CLI = Path("/Users/apple/Documents/Codex/2026-06-03/cli/lark-cli")
DRINK_DIR = ROOT / "阴山冲饮数据明细"
FOOD_DIR = ROOT / "阴山食品店数据明细"

REPORT = "Jos6sfYSRh4eWXtZalBcoFetnCe"
SCHEDULE = "WRHcsP1gYhSBhrtqNlvcnmPwnBh"
HANDOFF = "shtcnwOdgFZCQAf4ZjiR5egkoTc"
MONTHLY = "shtcnzv0loZQg1qk25PcCVonplW"
MONTHLY_SHEET = "zsta38"  # 2026-8 月度大表
LIVE_DATA = "shtcniNUd2HqPd8ZvZOt8btOScf"
LIVE_SHEET = "bea4e1"
LIVE_ACCOUNT = "阴山优麦冲饮旗舰店"
FOOD_SHOP_SHEET = "9VVamg"
FOOD_ACCOUNT = "阴山优麦食品旗舰店"
DRINK_LIVE_DETAIL_SHEET = "cC79qR"
FOOD_LIVE_DETAIL_SHEET = "weIxvN"

ACCOUNTS = [
    "阴山优麦冲饮旗舰店",
    "阴山优麦有机燕麦片",
    "阴山有机-宿州",
    "阴山优麦有机燕麦米",
    "蒙邮优选",
    "中国邮政乌兰察布分公司",
    "中国邮政集团有限公司乌兰察布市分公司",
]


def num(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def rounded(value, digits=2):
    return round(float(value) + 1e-9, digits)


def roi(gmv, cost):
    return rounded(gmv / cost, 2) if cost else 0


def excel_serial(day):
    return (day - dt.date(1899, 12, 30)).days


def col_name(index0):
    result = ""
    n = index0 + 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def run_lark_read(token, cell_range):
    cmd = [
        str(CLI),
        "sheets",
        "+read",
        "--as",
        "user",
        "--spreadsheet-token",
        token,
        "--range",
        cell_range,
        "--format",
        "json",
    ]
    proc = subprocess.run(cmd, check=True, text=True, capture_output=True)
    payload = json.loads(proc.stdout)
    return payload["data"]["valueRange"].get("values", [])


def lark_write_command(token, cell_range, values):
    return [
        str(CLI),
        "sheets",
        "+write",
        "--as",
        "user",
        "--spreadsheet-token",
        token,
        "--range",
        cell_range,
        "--values",
        json.dumps(values, ensure_ascii=False),
        "--format",
        "json",
    ]


def run_lark_write(token, cell_range, values):
    subprocess.run(lark_write_command(token, cell_range, values), check=True)


def shell_quote(text):
    return "'" + text.replace("'", "'\"'\"'") + "'"

def find_data_file(shop, prefix):
    """Find the target Excel in the shop's data folder, falling back to ROOT."""
    base = FOOD_DIR if shop == "food" else DRINK_DIR
    candidates = sorted(base.glob(f"{prefix}*.xlsx")) if base.exists() else []
    if not candidates:
        candidates = sorted(ROOT.glob(f"{prefix}*.xlsx"))
    return candidates[-1] if candidates else None


def load_live(day, shop="drink"):
    ymd = day.strftime("%Y%m%d")
    path = find_data_file(shop, f"直播明细_全部账号_{ymd}_{ymd}")
    if not path:
        print(f"Warning: live Excel for {ymd} not found, returning empty live data")
        return defaultdict(lambda: {"gmv": 0.0, "cost": 0.0})
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["直播间明细"]
    summary = defaultdict(lambda: {"gmv": 0.0, "cost": 0.0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name:
            continue
        summary[str(name)]["gmv"] += num(row[25])
        summary[str(name)]["cost"] += num(row[43])
    return summary


def load_live_detail(day, account_name, shop="drink"):
    """Load full detail rows for a specific account from the live stream Excel."""
    ymd = day.strftime("%Y%m%d")
    path = find_data_file(shop, f"直播明细_全部账号_{ymd}_{ymd}")
    if not path:
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["直播间明细"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == account_name:
            records.append(row)
    return records


def find_live_data_row(target_serial):
    """Find the row number in the live data sheet (bea4e1) matching the date serial."""
    values = run_lark_read(LIVE_DATA, f"{LIVE_SHEET}!A:A")
    for i, row in enumerate(values):
        if row and int(num(row[0])) == target_serial:
            return i + 1
    return None

def find_date_row(token, sheet, target_serial):
    """Find the row number in any sheet where column A matches an Excel serial."""
    values = run_lark_read(token, f"{sheet}!A:A")
    for i, row in enumerate(values):
        if row and int(num(row[0])) == target_serial:
            return i + 1
    return None


def aggregate_live_records(records):
    """Aggregate multiple live records: sum for volume metrics, max for rate/peak metrics."""
    return {
        "观看次数": sum(num(r[10]) for r in records),
        "成交金额": sum(num(r[25]) for r in records),
        "直播时长_min": sum(num(r[5]) for r in records),
        "投放消耗": sum(num(r[43]) for r in records),
        "成交件数": sum(num(r[28]) for r in records),
        "成交人数": sum(num(r[29]) for r in records),
        "最高在线": max(num(r[11]) for r in records),
        "平均在线": max(num(r[12]) for r in records),
        "平均停留_min": max(num(r[13]) for r in records),
        "观看成交率人数": max(num(r[39]) for r in records),
        "商品点击成交率人数": max(num(r[37]) for r in records),
    }





def load_compass(day, shop="drink"):
    ymd = day.strftime("%Y%m%d")
    path = find_data_file(shop, f"抖音电商罗盘-成交分析-{ymd}-{ymd}")
    if not path:
        print(f"Warning: compass Excel for {ymd} not found, returning empty compass data")
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["自营成交"]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0] if rows else []
    def hindex(name, fallback):
        for i, h in enumerate(header):
            if h is not None and str(h).strip() == name:
                return i
        return fallback
    amount_idx = hindex("成交金额", 3)
    user_idx = hindex("用户支付金额", amount_idx + 1)
    coupon_idx = hindex("智能优惠券金额", amount_idx + 2)
    subsidy_idx = hindex("平台补贴金额", amount_idx + 3)
    summary = {}
    for row in rows[1:]:
        date_str = str(row[0]) if row[0] is not None else ''
        if date_str != ymd:
            continue
        carrier = row[1] if len(row) > 1 else None
        slot = row[2] if len(row) > 2 else None
        if slot == "不限" and carrier in {"全部", "直播", "短视频", "商品卡", "图文", "其他"}:
            summary[carrier] = {
                "成交金额": num(row[amount_idx]),
                "用户支付金额": num(row[user_idx]),
                "智能优惠券金额": num(row[coupon_idx]),
                "平台补贴金额": num(row[subsidy_idx]),
            }
    return summary


def eval_formula(value, cells, visiting=None):
    if not isinstance(value, str):
        return num(value)
    expr = value.strip()
    if not expr:
        return 0.0
    if visiting is None:
        visiting = set()

    def replace_ref(match):
        ref = match.group(0)
        if ref in visiting:
            return "0"
        visiting.add(ref)
        resolved = eval_formula(cells.get(ref), cells, visiting)
        visiting.remove(ref)
        return str(resolved)

    expr = re.sub(r"\b[A-Z]{1,3}[0-9]+\b", replace_ref, expr)
    if not re.fullmatch(r"[0-9+\-*/(). ]+", expr):
        return 0.0
    tree = ast.parse(expr, mode="eval")
    for node in ast.walk(tree):
        if not isinstance(
            node,
            (
                ast.Expression,
                ast.BinOp,
                ast.UnaryOp,
                ast.Add,
                ast.Sub,
                ast.Mult,
                ast.Div,
                ast.USub,
                ast.UAdd,
                ast.Constant,
                ast.Load,
            ),
        ):
            return 0.0
    return float(eval(compile(tree, "<formula>", "eval"), {"__builtins__": {}}, {}))


def load_handoff(target_serial):
    values = run_lark_read(HANDOFF, "KFTIUP!A:I")
    cells = {}
    for r_index, row in enumerate(values, start=1):
        for c_index, value in enumerate(row):
            cells[f"{col_name(c_index)}{r_index}"] = value

    for i, row in enumerate(values):
        if row and int(num(row[0])) == target_serial:
            result = []
            for c in (2, 4, 6, 8):
                name = row[c] if c < len(row) else None
                if not name:
                    continue
                hours = num(values[i + 1][c]) if i + 1 < len(values) and c < len(values[i + 1]) else 0
                gmv_cell = f"{col_name(c)}{i + 3}"
                cost_cell = f"{col_name(c)}{i + 4}"
                gmv = rounded(eval_formula(cells.get(gmv_cell), cells), 2)
                cost = rounded(eval_formula(cells.get(cost_cell), cells), 2)
                result.append([str(name).replace("冲饮主播：", ""), hours, cost, gmv, roi(gmv, cost), rounded(gmv / hours, 2) if hours else 0])
            return result[:4]
    return []


def load_schedule(serial_value):
    values = run_lark_read(SCHEDULE, "3ce64b!A1:H500")
    for i, row in enumerate(values):
        if len(row) > 1 and int(num(row[1])) == serial_value:
            pairs = []
            for offset in range(2, 6):
                if i + offset < len(values):
                    r = values[i + offset]
                    pairs.append([r[2] if len(r) > 2 else "", r[3] if len(r) > 3 else ""])
            return pairs[:4]
    return [["", ""] for _ in range(4)]


def import_live_writes(day, shop):
    """Append target-date live detail rows to cC79qR (drink) or weIxvN (food)."""
    account = LIVE_ACCOUNT if shop == "drink" else FOOD_ACCOUNT
    sheet = DRINK_LIVE_DETAIL_SHEET if shop == "drink" else FOOD_LIVE_DETAIL_SHEET
    records = [list(r) for r in load_live_detail(day, account, shop)]
    if not records:
        return []
    existing = run_lark_read(REPORT, f"{sheet}!A1:D1000")
    def row_date(v):
        m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", str(v or "").strip())
        return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None
    for row in existing[1:]:
        if len(row) > 3 and str(row[1] or "").strip() == account and row_date(row[3]) == (day.year, day.month, day.day):
            return []
    values = run_lark_read(REPORT, f"{sheet}!A1:A1000")
    start = 2
    for i, row in enumerate(values, start=1):
        if not row or not row[0]:
            start = i
            break
    else:
        start = len(values) + 1
    end = start + len(records) - 1
    return [(REPORT, f"{sheet}!A{start}:BB{end}", records)]

def build_writes(day, offline_example=False, qianchuan_cost=None, shop="drink"):
    serial = excel_serial(day)
    day_num = day.day
    live = load_live(day, shop)
    compass = load_compass(day, shop)

    oat_flake = live.get("阴山有机-宿州") or live.get("阴山优麦有机燕麦片", {"gmv": 0, "cost": 0})
    oat_rice = live.get("阴山优麦有机燕麦米", {"gmv": 0, "cost": 0})
    postal = live.get("蒙邮优选") or live.get("中国邮政乌兰察布分公司") or live.get("中国邮政集团有限公司乌兰察布市分公司", {"gmv": 0, "cost": 0})
    total_cost = rounded(sum(live.get(name, {"cost": 0})["cost"] for name in ACCOUNTS), 2)

    carriers = {
        k: compass.get(k, {"成交金额": 0})["成交金额"]
        for k in ["直播", "短视频", "商品卡", "图文", "其他"]
    }
    report_total_gmv = rounded(sum(carriers.values()), 2)
    full_gmv = rounded(report_total_gmv - oat_flake["gmv"] - oat_rice["gmv"], 2)

    monthly_row = day_num + 4

    writes = []
    add = writes.append

    if shop == "food":
        food_row = find_date_row(REPORT, FOOD_SHOP_SHEET, serial)
        if food_row:
            add((REPORT, f"{FOOD_SHOP_SHEET}!B{food_row}", [[report_total_gmv]]))
            if qianchuan_cost is not None:
                add((REPORT, f"{FOOD_SHOP_SHEET}!C{food_row}", [[rounded(qianchuan_cost, 2)]]))
            add((REPORT, f"{FOOD_SHOP_SHEET}!D{food_row}", [[rounded(carriers["直播"])]]))
            add((REPORT, f"{FOOD_SHOP_SHEET}!E{food_row}", [[rounded(carriers["短视频"])]]))
            add((REPORT, f"{FOOD_SHOP_SHEET}!F{food_row}", [[rounded(carriers["商品卡"])]]))
            add((REPORT, f"{FOOD_SHOP_SHEET}!G{food_row}", [[rounded(carriers["其他"])]]))
            add((REPORT, f"{FOOD_SHOP_SHEET}!H{food_row}", [[rounded(carriers["图文"])]]))
            if qianchuan_cost is not None:
                add((REPORT, f"{FOOD_SHOP_SHEET}!I{food_row}", [[roi(report_total_gmv, qianchuan_cost)]]))
        add((MONTHLY, f"{MONTHLY_SHEET}!K{monthly_row}", [[report_total_gmv]]))
        if qianchuan_cost is not None:
            add((MONTHLY, f"{MONTHLY_SHEET}!M{monthly_row}", [[rounded(qianchuan_cost, 2)]]))
        writes.extend(import_live_writes(day, shop))
        return writes

    # UUAtO2 — 仪表盘数据源
    uu_row = day_num + 6
    # 当前 UUAtO2 表头在第 1 行，8月1日从第 2 行开始
    uu_row = day_num + 1
    add((REPORT, f"UUAtO2!B{uu_row}", [[report_total_gmv]]))
    if qianchuan_cost is not None:
        add((REPORT, f"UUAtO2!C{uu_row}", [[rounded(qianchuan_cost, 2)]]))
    # 千川消耗（C列）由人工填写，脚本不再写入（2026-08-19 用户确认）
    add((REPORT, f"UUAtO2!D{uu_row}", [[rounded(carriers["直播"])]]))
    add((REPORT, f"UUAtO2!E{uu_row}", [[rounded(carriers["短视频"])]]))
    add((REPORT, f"UUAtO2!F{uu_row}", [[rounded(carriers["商品卡"])]]))
    add((REPORT, f"UUAtO2!G{uu_row}", [[rounded(carriers["其他"])]]))
    add((REPORT, f"UUAtO2!H{uu_row}", [[rounded(carriers["图文"])]]))
    add((REPORT, f"UUAtO2!J{uu_row}", [[rounded(postal["gmv"])]]))
    add((REPORT, f"UUAtO2!K{uu_row}", [[rounded(postal["cost"])]]))
    add((REPORT, f"UUAtO2!L{uu_row}", [[roi(postal["gmv"], postal["cost"])]]))

    add((MONTHLY, f"{MONTHLY_SHEET}!D{monthly_row}", [[full_gmv]]))
    if qianchuan_cost is not None:
        add((MONTHLY, f"{MONTHLY_SHEET}!F{monthly_row}", [[rounded(qianchuan_cost, 2)]]))
    for col, data in [("Q", postal)]:
        gmv = rounded(data["gmv"])
        cost = rounded(data["cost"])
        # Convert by index instead of relying on ASCII after Z.
        start_index = 0
        for ch in col:
            start_index = start_index * 26 + ord(ch) - 64
        start_index -= 1
        add((MONTHLY, f"{MONTHLY_SHEET}!{col}{monthly_row}", [[gmv]]))
        add((MONTHLY, f"{MONTHLY_SHEET}!{col_name(start_index + 1)}{monthly_row}", [[cost]]))
        add((MONTHLY, f"{MONTHLY_SHEET}!{col_name(start_index + 2)}{monthly_row}", [[roi(gmv, cost)]]))

    # 阴山优麦冲饮旗舰店直播间数据记录表
    records = load_live_detail(day, LIVE_ACCOUNT, shop)
    if records:
        agg = aggregate_live_records(records)
        live_row = find_live_data_row(serial)
        if live_row:
            duration_h = round(agg["直播时长_min"] / 60)
            avg_stay_sec = round(agg["平均停留_min"] * 60, 2)
            add((LIVE_DATA, f"{LIVE_SHEET}!B{live_row}", [[round(agg["观看次数"])]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!C{live_row}", [[round(agg["成交金额"], 2)]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!E{live_row}", [[duration_h]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!F{live_row}", [[round(agg["投放消耗"], 2)]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!G{live_row}", [[round(agg["成交金额"], 2)]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!H{live_row}", [[round(agg["最高在线"])]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!I{live_row}", [[round(agg["平均在线"])]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!J{live_row}", [[avg_stay_sec]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!K{live_row}", [[round(agg["成交件数"])]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!L{live_row}", [[round(agg["成交人数"])]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!M{live_row}", [[agg["观看成交率人数"]]]))
            add((LIVE_DATA, f"{LIVE_SHEET}!N{live_row}", [[agg["商品点击成交率人数"]]]))

    writes.extend(import_live_writes(day, shop))
    return writes


def emit_script(writes, output):
    lines = ["#!/usr/bin/env zsh", "set -euo pipefail", ""]
    for token, cell_range, values in writes:
        cmd = lark_write_command(token, cell_range, values)
        lines.append(" ".join(shell_quote(part) for part in cmd))
    lines.append('echo "Done. Read back Feishu sheets to verify."')
    Path(output).write_text("\n".join(lines) + "\n", encoding="utf-8")
    Path(output).chmod(0o755)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="Target date, YYYY-MM-DD. Defaults to yesterday.")
    parser.add_argument("--mode", choices=["plan", "write", "script"], default="plan")
    parser.add_argument("--output", default=str(ROOT / "write_yinshan_report.sh"))
    parser.add_argument("--offline-example", action="store_true", help="Use embedded 2026-06-03 Feishu schedule/handoff sample for offline self-test.")
    parser.add_argument("--qianchuan-cost", type=float, help="Qianchuan account total cost for the target date.")
    parser.add_argument("--shop", choices=["drink", "food"], default="drink", help="Shop to fill: drink (冲饮店) or food (食品店).")
    args = parser.parse_args()

    day = dt.date.fromisoformat(args.date) if args.date else dt.date.today() - dt.timedelta(days=1)
    writes = build_writes(day, offline_example=args.offline_example, qianchuan_cost=args.qianchuan_cost, shop=args.shop)
    if args.mode == "plan":
        print(json.dumps([{"token": t, "range": r, "values": v} for t, r, v in writes], ensure_ascii=False, indent=2))
    elif args.mode == "write":
        for token, cell_range, values in writes:
            run_lark_write(token, cell_range, values)
        print("Done. Read back Feishu sheets to verify.")
        # 2026-08-19: 写入完成后自动刷新驾驶舱（dashboard_data_updater.py）
        updater = ROOT / "dashboard_data_updater.py"
        if updater.exists():
            print("[yinshan_daily_report] Refreshing dashboard data...")
            proc = subprocess.run(
                [sys.executable, str(updater)], capture_output=True, text=True, timeout=120
            )
            if proc.stdout:
                print(proc.stdout.strip())
            if proc.returncode != 0:
                print(f"[yinshan_daily_report] WARNING: dashboard refresh failed (rc={proc.returncode})")
                if proc.stderr:
                    print(proc.stderr.strip()[-1500:])
            else:
                print("[yinshan_daily_report] Dashboard refreshed.")
        else:
            print(f"[yinshan_daily_report] WARNING: {updater} not found, skipped dashboard refresh")
    else:
        emit_script(writes, args.output)
        print(args.output)


if __name__ == "__main__":
    main()
